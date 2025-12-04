import face_recognition
from openpyxl import Workbook, load_workbook
import os
import glob
import numpy as np
import datetime

# ----------------- USER TUNABLE PARAMETERS -----------------
HOG_MODEL = "hog"
UPSAMPLE_TIMES = 2
IOU_DUPLICATE_THRESHOLD = 0.35
SMALL_FACE_HEIGHT_PX = 60
CONFIDENCE_THRESHOLD_CLOSE = 55.0
CONFIDENCE_THRESHOLD_SMALL = 55.0
COMPARE_TOLERANCE_CLOSE = 0.6
COMPARE_TOLERANCE_SMALL = 0.55
# -----------------------------------------------------------

# Attendance folder
ATTENDANCE_FOLDER = "Attendance"
if not os.path.exists(ATTENDANCE_FOLDER):
    os.makedirs(ATTENDANCE_FOLDER)

# -------------------- HELPER FUNCTIONS --------------------
def load_known_faces(dataset_folder="dataset"):
    known_face_encodings = []
    known_face_names = []

    if not os.path.exists(dataset_folder):
        os.makedirs(dataset_folder)
        print(f"Created dataset folder: {dataset_folder}. Add face images named by ID.")
        return known_face_encodings, known_face_names

    image_extensions = ["jpg", "jpeg", "png", "bmp", "tiff"]
    all_image_files = []
    for ext in image_extensions:
        all_image_files.extend(glob.glob(os.path.join(dataset_folder, f"*.{ext}")))
        all_image_files.extend(glob.glob(os.path.join(dataset_folder, f"*.{ext.upper()}")))

    all_image_files = list(set(all_image_files))
    for image_path in all_image_files:
        name = os.path.splitext(os.path.basename(image_path))[0]
        try:
            img = face_recognition.load_image_file(image_path)
            encs = face_recognition.face_encodings(img)
            if len(encs) > 0:
                known_face_encodings.append(encs[0])
                known_face_names.append(name)
        except Exception as e:
            print(f"Error loading {image_path}: {e}")

    return known_face_encodings, known_face_names


def iou(boxA, boxB):
    tA, rA, bA, lA = boxA
    tB, rB, bB, lB = boxB
    xI1 = max(lA, lB)
    yI1 = max(tA, tB)
    xI2 = min(rA, rB)
    yI2 = min(bA, bB)
    interW = max(0, xI2 - xI1)
    interH = max(0, yI2 - yI1)
    interArea = interW * interH
    areaA = (rA - lA) * (bA - tA)
    areaB = (rB - lB) * (bB - tB)
    union = areaA + areaB - interArea
    return interArea / union if union != 0 else 0.0


def merge_locations_iou(locations, iou_threshold=IOU_DUPLICATE_THRESHOLD):
    kept = []
    for loc in locations:
        if all(iou(loc, existing) <= iou_threshold for existing in kept):
            kept.append(loc)
    return kept


# ------------------ MAIN PROCESSING FUNCTION ------------------
def process_uploaded_image(image_path, known_face_encodings, known_face_names):
    """
    Returns:
        recognized_faces: list of dicts {'name','confidence','location'}
    """
    uploaded_image = face_recognition.load_image_file(image_path)
    all_locations = []

    # Strategy 1: fast HOG detection
    locs1 = face_recognition.face_locations(uploaded_image, model=HOG_MODEL)
    all_locations.extend(locs1)

    # Strategy 2: upsample pass
    locs2 = face_recognition.face_locations(uploaded_image, number_of_times_to_upsample=UPSAMPLE_TIMES, model=HOG_MODEL)
    for loc in locs2:
        if all(iou(loc, existing) <= IOU_DUPLICATE_THRESHOLD for existing in all_locations):
            all_locations.append(loc)

    unique_locations = merge_locations_iou(all_locations)
    face_encodings = face_recognition.face_encodings(uploaded_image, unique_locations)
    recognized_faces = []

    for i, face_encoding in enumerate(face_encodings):
        loc = unique_locations[i]
        top, right, bottom, left = loc
        face_height = bottom - top

        if face_height < SMALL_FACE_HEIGHT_PX:
            compare_tolerance = COMPARE_TOLERANCE_SMALL
            min_confidence = CONFIDENCE_THRESHOLD_SMALL
        else:
            compare_tolerance = COMPARE_TOLERANCE_CLOSE
            min_confidence = CONFIDENCE_THRESHOLD_CLOSE

        matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=compare_tolerance)
        name = "Unknown"
        confidence = 0.0
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        if len(face_distances) > 0:
            best_idx = np.argmin(face_distances)
            if matches[best_idx]:
                confidence = (1 - face_distances[best_idx]) * 100
                name = known_face_names[best_idx]

        recognized_faces.append({
            'name': name,
            'confidence': confidence,
            'location': loc
        })

    return recognized_faces


# ------------------ EXCEL FUNCTIONS ------------------
def save_attendance_excel(present_ids, known_face_names):
    now = datetime.datetime.now()
    today = now.day
    month = now.month
    year = now.year
    filename = os.path.join(ATTENDANCE_FOLDER, f"{year}-{month:02d}-{today:02d}.xlsx")

    if os.path.exists(filename):
        book = load_workbook(filename)
        sheet = book.active
    else:
        book = Workbook()
        sheet = book.active
        sheet.title = "Attendance"

    # Write present student IDs only
    present_ids_sorted = sorted(present_ids)
    for i, sid in enumerate(present_ids_sorted, start=2):  # row 2 onwards
        sheet.cell(row=i, column=1).value = sid

    # Summary
    total_present = len(present_ids_sorted)
    summary_start_row = len(present_ids_sorted) + 3
    sheet.cell(row=summary_start_row, column=7).value = "SUMMARY"
    sheet.cell(row=summary_start_row + 1, column=7).value = "Total Students"
    sheet.cell(row=summary_start_row + 1, column=8).value = len(known_face_names)
    sheet.cell(row=summary_start_row + 2, column=7).value = "Present Today"
    sheet.cell(row=summary_start_row + 2, column=8).value = total_present
    sheet.cell(row=summary_start_row + 3, column=7).value = "Absent Today"
    sheet.cell(row=summary_start_row + 3, column=8).value = len(known_face_names) - total_present
    sheet.cell(row=summary_start_row + 4, column=7).value = "Attendance Rate"
    sheet.cell(row=summary_start_row + 4, column=8).value = f"{(total_present / len(known_face_names) * 100):.1f}%" if known_face_names else "0%"
    sheet.cell(row=summary_start_row + 6, column=7).value = "Report Generated"
    sheet.cell(row=summary_start_row + 6, column=8).value = now.strftime("%Y-%m-%d %H:%M:%S")

    book.save(filename)
    return filename


# ------------------ BACKEND FUNCTION ------------------
def get_attendance_from_image(image_path, save_excel=True):
    """
    Stateless backend function.
    Returns:
        marked_ids: sorted list of students who were actually marked present
        recognized_faces: full info for each detected face
    """
    # Load known faces
    known_face_encodings, known_face_names = load_known_faces("dataset")
    if not known_face_encodings:
        return [], []

    # Process the image
    recognized_faces = process_uploaded_image(image_path, known_face_encodings, known_face_names)

    # Filter only faces that pass confidence threshold (marked attendance)
    marked_ids = []
    for f in recognized_faces:
        # Only mark known students with confidence > threshold
        if f['name'] != "Unknown":
            # Determine face height from location
            top, right, bottom, left = f['location']
            face_height = bottom - top
            if face_height < SMALL_FACE_HEIGHT_PX:
                min_confidence = CONFIDENCE_THRESHOLD_SMALL
            else:
                min_confidence = CONFIDENCE_THRESHOLD_CLOSE

            if f['confidence'] > min_confidence:
                marked_ids.append(f['name'])

    # Remove duplicates and sort
    marked_ids = sorted(list(set(marked_ids)))

    # Save to Excel if requested
    if save_excel and marked_ids:
        save_attendance_excel(marked_ids, known_face_names)

    return marked_ids, recognized_faces



# ------------------ TEST ------------------
if __name__ == "__main__":
    test_image = "D:\pinterest\photo_5859528924960852754_y.jpg"  # replace with your image path
    present_ids, recognized_faces = get_attendance_from_image(test_image)
    print("Present student IDs:", present_ids)
